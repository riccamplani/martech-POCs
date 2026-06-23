"""Write the field-usage rollup to an Excel workbook matching the project's
7-group sheet structure (Account, Account_Structure_Summary, ... PS_User_TS_S_EU),
plus a Summary sheet with COUNTIF-based totals.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .aggregate import ObjectRollup

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
UNUSED_FILL = PatternFill("solid", start_color="FCE4E4")


def write_workbook(rollups: dict[str, ObjectRollup], out_path: Path) -> Path:
    by_group: dict[str, list[ObjectRollup]] = defaultdict(list)
    for roll in rollups.values():
        by_group[roll.group].append(roll)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    # one sheet per analysis group
    for group, rolls in by_group.items():
        ws = wb.create_sheet(title=group[:31])  # Excel 31-char sheet name cap
        ws.append(["Object", "Field", "Status", "# Recipes Using", "Example Recipe"])
        for c in ws[1]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        for roll in sorted(rolls, key=lambda r: r.object_name):
            for fld in sorted(roll.loaded):
                used = fld in roll.used
                recipes = roll.used_by.get(fld, [])
                row = [roll.object_name, fld, "USED" if used else "UNUSED",
                       len(recipes), recipes[0] if recipes else ""]
                ws.append(row)
                if not used:
                    for cell in ws[ws.max_row]:
                        cell.fill = UNUSED_FILL
        for col, width in zip("ABCDE", (34, 40, 10, 16, 45)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        for r in ws.iter_rows(min_row=1):
            for cell in r:
                if cell.font is not HEADER_FONT:
                    cell.font = Font(name=FONT)

    # summary sheet: per-object counts via COUNTIF over each group sheet
    summary.append(["Group", "Object", "Loaded", "Used", "Unused", "% Unused"])
    for c in summary[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for group, rolls in by_group.items():
        sheet = group[:31]
        for roll in sorted(rolls, key=lambda r: r.object_name):
            obj = roll.object_name
            loaded = len(roll.loaded)
            r = summary.max_row + 1
            summary.append([
                group, obj,
                f'=COUNTIFS(\'{sheet}\'!A:A,B{r})',
                f'=COUNTIFS(\'{sheet}\'!A:A,B{r},\'{sheet}\'!C:C,"USED")',
                f'=COUNTIFS(\'{sheet}\'!A:A,B{r},\'{sheet}\'!C:C,"UNUSED")',
                f'=IF(C{r}=0,0,E{r}/C{r})',
            ])
            summary.cell(row=r, column=6).number_format = "0.0%"
    for col, width in zip("ABCDEF", (26, 34, 10, 8, 9, 10)):
        summary.column_dimensions[col].width = width
    summary.freeze_panes = "A2"
    for r in summary.iter_rows(min_row=2):
        for cell in r:
            cell.font = Font(name=FONT)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
